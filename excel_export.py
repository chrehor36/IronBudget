# -*- coding: utf-8 -*-
"""IronBudget Excel export - the same plain-styled workbook the console tool
built, adapted to take pre-computed data and a destination path instead of
building its own data and picking its own filename."""
import datetime as dt
import os
import time

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList

CUR, CUR0, PCT = '"$"#,##0.00', '"$"#,##0', '0.0%'
HEADER_FILL = "D9D9D9"


def F(sz=10, b=False, i=False):
    return Font(name="Calibri", size=sz, bold=b, italic=i)


def fillc(h):
    return PatternFill("solid", fgColor=h)


def _title(ws, text, sub, cols=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = F(16, True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    c2 = ws.cell(row=2, column=1, value=sub)
    c2.font = F(9, i=True)
    ws.row_dimensions[1].height = 22


def _table(ws, box, r0, col0, headers, data, widths, fmts, total=None):
    for i, h in enumerate(headers):
        c = ws.cell(row=r0, column=col0 + i, value=h)
        c.font = F(10, True); c.fill = fillc(HEADER_FILL); c.border = box
        c.alignment = Alignment(horizontal="left", vertical="center")
    r = r0 + 1
    for row in data:
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=col0 + i - 1, value=v)
            c.font = F(10); c.border = box
            if i in fmts and isinstance(v, (int, float)):
                c.number_format = fmts[i]
        r += 1
    if total:
        for i, v in enumerate(total, 1):
            c = ws.cell(row=r, column=col0 + i - 1, value=v)
            c.font = F(10, True); c.border = box
            if i in fmts and isinstance(v, (int, float)):
                c.number_format = fmts[i]
        r += 1
    for i, wd in enumerate(widths):
        ws.column_dimensions[get_column_letter(col0 + i)].width = wd
    return r


def build_workbook(agg, rows, budget_title, dest_path):
    """agg: the dict returned by engine.compute_aggregates (native, not JSON-safe).
    rows: the native transaction list (dt.date objects intact).
    budget_title: e.g. "Chris Hrehor Budget".
    dest_path: full .xlsx path chosen by the user via the Save As dialog.

    Pure function: data + path in, file on disk out. Does not open the file
    afterward - that's the caller's job."""
    cat_sorted = agg["cat_sorted"]
    accounts = agg["accounts"]
    has_transfers = agg["has_transfers"]
    xfer_m = agg["xfer_m"]; xfer_total = agg["xfer_total"]
    exp_adj = agg["exp_adj"]; net_m = agg["net_m"]
    detected_trips = agg["detected_trips"]; home_state = agg["home_state"]
    trip_cost = agg["trip_cost"]; trip_reimb = agg["trip_reimb"]
    lumpy = agg["lumpy"]
    m_inc = agg["m_inc"]; m_exp_adj = agg["m_exp_adj"]; m_xfer = agg["m_xfer"]; months = agg["months"]
    inc_total = agg["inc_total"]; MONTHS = agg["MONTHS"]; exp_m = agg["exp_m"]
    w_raw = agg["w_raw"]; w_adj = agg["w_adj"]; w_inc = agg["w_inc"]
    recurring = agg["recurring"]; xfer_by_acct = agg["xfer_by_acct"]
    START = agg["START"]; END = agg["END"]; DAYS = agg["DAYS"]
    inc_m = agg["inc_m"]

    wb = Workbook()
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    def table(ws, r0, col0, headers, data, widths, fmts, total=None):
        return _table(ws, box, r0, col0, headers, data, widths, fmts, total)

    def title(ws, text, sub, cols=10):
        _title(ws, text, sub, cols)

    # ---------- Dashboard ----------
    ws = wb.active; ws.title = "Dashboard"
    acct_note = f"{len(accounts)} accounts: {', '.join(accounts)}" if len(accounts) > 1 else f"{accounts[0]} account"
    title(ws, budget_title, f"{START:%b %d, %Y} - {END:%b %d, %Y}  ({DAYS} days, ~{MONTHS:.1f} months)  |  {acct_note}", 10)

    kpis = [
        ("Total income / mo", inc_m, "Paychecks, refunds, interest"),
        ("Total expenses / mo", exp_m, "Adjusted spend"),
        ("Surplus / mo", net_m, "Income minus expenses"),
    ]
    if has_transfers:
        kpis.append(("Moved between accounts / mo", xfer_m, f"${xfer_total:,.0f} total across {len(accounts)} accounts"))
    r = table(ws, 4, 1, ["METRIC", "AMOUNT", "NOTE"], [[l, v, s] for l, v, s in kpis], [30, 15, 48], {2: CUR0})

    r += 1
    ws.cell(row=r, column=1, value="Where the money goes").font = F(12, True)
    r += 1
    cat_r0 = r
    cdata = [[c, v, v / exp_adj if exp_adj else 0] for c, v in cat_sorted]
    last_cat_row = table(ws, cat_r0, 1, ["CATEGORY", "AMOUNT", "% OF SPEND"], cdata, [26, 14, 12],
                         {2: CUR0, 3: PCT}, total=["TOTAL", exp_adj, 1.0])

    topn = cat_sorted[:7]
    other_amt = sum(v for _, v in cat_sorted[7:])
    donut_rows = topn + ([("Other", other_amt)] if other_amt > 0 else [])
    d0 = cat_r0
    for i, (c, v) in enumerate(donut_rows):
        ws.cell(row=d0 + i, column=10, value=c)
        ws.cell(row=d0 + i, column=11, value=v)
    d_last = d0 + len(donut_rows) - 1
    donut = DoughnutChart()
    donut.title = "Where the money goes"
    donut.add_data(Reference(ws, min_col=11, min_row=d0, max_row=d_last), titles_from_data=False)
    donut.set_categories(Reference(ws, min_col=10, min_row=d0, max_row=d_last))
    donut.height, donut.width = 9, 11
    donut.visible_cells_only = False
    donut.series[0].dLbls = DataLabelList()
    donut.series[0].dLbls.showSerName = False
    donut.series[0].dLbls.showCatName = False
    donut.series[0].dLbls.showVal = False
    donut.series[0].dLbls.showLegendKey = False
    donut.series[0].dLbls.showPercent = True
    ws.add_chart(donut, f"E{cat_r0}")
    for hide_col in ("J", "K"):
        ws.column_dimensions[hide_col].hidden = True

    r2 = last_cat_row + 1
    gap = net_m - xfer_m
    if gap >= 0 or not has_transfers:
        banner = f"Income covers expenses with ${net_m:,.0f}/mo to spare."
    else:
        banner = f"Bills covered with ${net_m:,.0f}/mo to spare, but your savings pace runs ${-gap:,.0f}/mo ahead of that surplus."
    ws.cell(row=r2, column=1, value=banner).font = F(11, True)
    r2 += 2

    if detected_trips:
        ws.cell(row=r2, column=1, value=f"Auto-detected trips (home base: {home_state})").font = F(11, True); r2 += 1
        trip_data = []
        for d in detected_trips:
            span = f"{d['start']:%b %d} - {d['end']:%b %d}" if d['start'] != d['end'] else f"{d['start']:%b %d}"
            where = ", ".join(d["locations"])
            reimb_note = f"${d['reimb_amt']:,.0f} on {d['reimb_date']:%b %d}" if d["reimb_amt"] else "none found — counted as a real expense"
            trip_data.append([span, where, d["cost"], reimb_note])
        r2 = table(ws, r2, 1, ["DATES", "STATE(S)", "COST", "REIMBURSEMENT"], trip_data,
                   [20, 14, 12, 34], {3: CUR}, total=["TOTAL", "", trip_cost, ""]) + 1
        net_oop = trip_cost - trip_reimb
        ws.cell(row=r2, column=1, value=f"Net out of pocket across all detected trips: ${net_oop:,.0f}").font = F(9, i=True)
        r2 += 2
    elif trip_cost or trip_reimb:
        ws.cell(row=r2, column=1, value="Trip reconciliation").font = F(11, True); r2 += 1
        r2 = table(ws, r2, 1, ["ITEM", "AMOUNT"],
                   [["Trip expenses fronted", -trip_cost], ["Reimbursement received", trip_reimb],
                    ["Net out of pocket", trip_reimb - trip_cost]], [30, 14], {2: CUR}) + 1

    ws.cell(row=r2, column=1, value="Notable one-time purchases (>= $100)").font = F(11, True); r2 += 1
    table(ws, r2, 1, ["PURCHASE", "AMOUNT"],
          [[f"{t['date']:%b %d} - {t['desc']} ({t['cat']})", -t["amt"]] for t in lumpy][:14],
          [50, 14], {2: CUR})

    # ---------- Monthly ----------
    ws = wb.create_sheet("Monthly")
    title(ws, "Monthly Cash Flow", "Adjusted = detected/configured trip removed. Transfers between your own accounts shown separately.", 6)
    data = [[dt.date.fromisoformat(k + "-01").strftime("%b %Y"), m_inc[k], m_exp_adj[k],
             m_inc[k] - m_exp_adj[k], m_xfer[k]] for k in months]
    r = table(ws, 4, 1, ["MONTH", "INCOME", "EXPENSES", "NET", "TRANSFERRED OUT"], data,
              [12, 14, 14, 13, 16], {2: CUR0, 3: CUR0, 4: CUR0, 5: CUR0},
              total=["TOTAL", inc_total, exp_adj, inc_total - exp_adj, xfer_total])
    n = len(months)
    ch = BarChart(); ch.type = "col"; ch.title = "Income vs expenses by month"
    ch.add_data(Reference(ws, min_col=2, min_row=4, max_row=4 + n), titles_from_data=True)
    ch.add_data(Reference(ws, min_col=3, min_row=4, max_row=4 + n), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4 + n))
    ch.gapWidth = 60; ch.y_axis.numFmt = CUR0
    ch.height, ch.width = 9.5, 21
    ws.add_chart(ch, f"A{r + 2}")
    ch2 = LineChart(); ch2.title = "Net cash flow by month"
    ch2.add_data(Reference(ws, min_col=4, min_row=4, max_row=4 + n), titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4 + n))
    ch2.y_axis.numFmt = CUR0
    ch2.height, ch2.width = 9.5, 21
    ws.add_chart(ch2, f"A{r + 22}")

    # ---------- Categories ----------
    ws = wb.create_sheet("Categories")
    title(ws, "Where the money goes", "Adjusted spend by category, full period.", 6)
    table(ws, 4, 1, ["CATEGORY", "TOTAL", "PER MONTH", "% OF SPEND"],
          [[c, v, v / MONTHS, v / exp_adj if exp_adj else 0] for c, v in cat_sorted], [26, 14, 13, 12],
          {2: CUR, 3: CUR, 4: PCT}, total=["TOTAL", exp_adj, exp_m, 1.0])
    bar_top = cat_sorted[:12]
    bar_other = sum(v for _, v in cat_sorted[12:])
    bar_rows = bar_top + ([("Other", bar_other)] if bar_other > 0 else [])
    b0 = 5
    for i, (c, v) in enumerate(reversed(bar_rows)):
        ws.cell(row=b0 + i, column=10, value=c)
        ws.cell(row=b0 + i, column=11, value=v)
    b_last = b0 + len(bar_rows) - 1
    ch = BarChart(); ch.type = "bar"; ch.title = "Spend by category (top 12 + Other)"
    ch.add_data(Reference(ws, min_col=11, min_row=b0, max_row=b_last), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=10, min_row=b0, max_row=b_last))
    ch.legend = None; ch.x_axis.numFmt = CUR0
    ch.visible_cells_only = False
    ch.height, ch.width = 13.5, 20
    ws.add_chart(ch, "G4")
    for cn in ("J", "K"):
        ws.column_dimensions[cn].hidden = True

    # ---------- Trends ----------
    ws = wb.create_sheet("Trends")
    title(ws, "Spending trends", "Weekly spend (Mon-start weeks) and cumulative surplus.", 6)
    allw = sorted(set(w_raw) | set(w_adj) | set(w_inc))
    data, cum = [], 0.0
    for wkk in allw:
        cum += w_inc.get(wkk, 0) - w_adj.get(wkk, 0)
        data.append([wkk.strftime("%b %d"), w_raw.get(wkk, 0.0), w_adj.get(wkk, 0.0), cum])
    table(ws, 4, 1, ["WEEK OF", "SPEND (RAW)", "SPEND (ADJ)", "CUMULATIVE SURPLUS"], data,
          [11, 13, 13, 18], {2: CUR0, 3: CUR0, 4: CUR0})
    nw = len(data)
    label_skip = max(1, nw // 12)
    ch = LineChart(); ch.title = "Weekly spending: raw vs adjusted"
    ch.add_data(Reference(ws, min_col=2, min_row=4, max_col=3, max_row=4 + nw), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4 + nw))
    ch.y_axis.numFmt = CUR0
    ch.x_axis.tickLblSkip = label_skip
    ch.height, ch.width = 9.5, 22
    ws.add_chart(ch, "F4")
    ch2 = LineChart(); ch2.title = "Cumulative surplus"
    ch2.add_data(Reference(ws, min_col=4, min_row=4, max_row=4 + nw), titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4 + nw))
    ch2.y_axis.numFmt = CUR0
    ch2.x_axis.tickLblSkip = label_skip
    ch2.height, ch2.width = 9.5, 22
    ws.add_chart(ch2, "F24")

    # ---------- Transfers (only if present) ----------
    if has_transfers:
        ws = wb.create_sheet("Transfers")
        title(ws, "Transfers between your own accounts", f"Across {len(accounts)} detected accounts: {', '.join(accounts)}. Not counted as spending.", 6)
        mdata = [[dt.date.fromisoformat(k + "-01").strftime("%b %Y"), m_xfer[k]] for k in months]
        r = table(ws, 4, 1, ["MONTH", "TRANSFERRED OUT (all accounts)"], mdata, [16, 24], {2: CUR0},
                  total=["TOTAL", xfer_total])
        ch = BarChart(); ch.type = "col"; ch.title = "Transfers by month"
        ch.add_data(Reference(ws, min_col=2, min_row=4, max_row=4 + len(mdata)), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4 + len(mdata)))
        ch.gapWidth = 60; ch.y_axis.numFmt = CUR0; ch.legend = None
        ch.height, ch.width = 9.5, 18
        ws.add_chart(ch, "D4")

        acct_data = sorted(xfer_by_acct.items(), key=lambda kv: -kv[1])
        table(ws, r + 2, 1, ["ACCOUNT (money left FROM here)", "TOTAL SENT OUT"],
              [[a, v] for a, v in acct_data], [30, 16], {2: CUR0})

    # ---------- Recurring ----------
    ws = wb.create_sheet("Recurring")
    title(ws, "Recurring charges (auto-detected)", "Merchants seen 2+ times. 'Est. monthly' = total / months elapsed. Review before trusting fully.", 6)
    data, sub_total = [], 0.0
    for label, ts in recurring[:25]:
        tot = sum(-t["amt"] for t in ts); sub_total += tot
        data.append([label, len(ts), tot, tot / MONTHS])
    table(ws, 4, 1, ["ITEM", "HITS", "TOTAL", "EST. MONTHLY"], data, [30, 7, 14, 14],
          {3: CUR, 4: CUR}, total=["TOTAL (shown)", "", sub_total, sub_total / MONTHS])

    # ---------- Transactions ----------
    ws = wb.create_sheet("Transactions")
    title(ws, "All transactions (repeats removed)", "Flags: Trip / Reimb apply automatically, or from the optional CONFIG block. Internal = between your own accounts.", 7)
    hdr = ["DATE", "ACCOUNT", "DESCRIPTION", "CATEGORY", "AMOUNT", "STATUS", "FLAGS"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = F(10, True); c.fill = fillc(HEADER_FILL); c.border = box
    r = 5
    for t in sorted(rows, key=lambda x: (x["date"], x["acct"]), reverse=True):
        flags = " ".join(f for f, b in [("Trip", t["trip"]), ("Reimb", t["reimb"]), ("Internal", t["internal"])] if b)
        for i, v in enumerate([t["date"], t["acct"], t["desc"], t["cat"], t["amt"], t["status"], flags], 1):
            c = ws.cell(row=r, column=i, value=v); c.font = F(10); c.border = box
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=5).number_format = CUR
        r += 1
    ws.auto_filter.ref = f"A4:G{r - 1}"
    ws.freeze_panes = "A5"
    for i, wd in enumerate([11, 10, 36, 24, 12, 9, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    # ------------------------- save (retry if locked) -------------------------
    for attempt in range(1, 6):
        try:
            wb.save(dest_path)
            return
        except PermissionError:
            time.sleep(2)
    raise PermissionError(f"Could not save - '{os.path.basename(dest_path)}' is open in another program. Close it and try again.")
