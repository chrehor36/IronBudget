# -*- coding: utf-8 -*-
"""
IronBudget — turns bank-export CSVs in this folder into "IronBudget.xlsx".

HOW TO USE
  1. Export transaction CSVs from your bank/finance app into this same folder.
     - Any filename works. Put the word "saving" in the filename for a
       savings-account export; everything else is treated as checking.
     - Overlapping exports are fine — duplicates are removed automatically.
  2. Run this file (double-click the .exe, or `python BudgetBuilder.py`).
  3. "IronBudget.xlsx" is built (and opened) in this same folder.

Expected CSV columns: Date, Description, Original Description, Category, Amount, Status

TRIPS ARE DETECTED AUTOMATICALLY: most bank exports embed a merchant's city/state
in the Original Description (e.g. "...CHICAGO      IL"). This tool figures out
your home state (whichever state shows up most often) and flags any cluster of
away-from-home spending as a trip, then tries to match it to a reimbursement
deposit that lands within ~45 days afterward. No dates to enter. You can still
add manual trips/reimbursement rules in the optional CONFIG block below if the
auto-detection misses something or you want to force a specific window.
"""
import csv, glob, json, os, re, sys, time, datetime as dt
from collections import defaultdict, Counter

try:
    import openpyxl  # noqa: F401
except ImportError:
    import subprocess
    print("First run: installing a required library (openpyxl)...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"])

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList

FOLDER = os.path.dirname(os.path.abspath(sys.argv[0]))

# ------------------------- who is this budget for? -------------------------
def ask(prompt):
    try:
        return input(prompt).strip()
    except EOFError:
        return ""

def join_and(items):
    items = list(items)
    if len(items) == 1:
        return items[0]
    if len(items) == 2:
        return f"{items[0]} & {items[1]}"
    return ", ".join(items[:-1]) + f" & {items[-1]}"

print("=" * 44)
print("  IRONBUDGET")
print("=" * 44)
while True:
    n_raw = ask("How many people are part of this budget? ")
    if n_raw.isdigit() and int(n_raw) >= 1:
        n_people = int(n_raw)
        break
    print("Please enter a whole number, 1 or more.")

people = []
for i in range(1, n_people + 1):
    first = ask(f"  Person {i} - first name: ") or f"Person{i}"
    last = ask(f"  Person {i} - last name: ")
    people.append((first, last))

if n_people == 1:
    first, last = people[0]
    budget_title = f"{first} {last} Budget".replace("  ", " ").strip()
else:
    lasts_seen = []
    for _, last in people:
        if last and last not in lasts_seen:
            lasts_seen.append(last)
    if len(lasts_seen) == 1:
        budget_title = f"{join_and([p[0] for p in people])} {lasts_seen[0]} Budget"
    elif lasts_seen:
        budget_title = f"{join_and(lasts_seen)} Budget"
    else:
        budget_title = "Household Budget"

print(f"\nBuilding: {budget_title}\n")
SAFE_TITLE = re.sub(r'[\\/:*?"<>|]', "", budget_title)
OUT = os.path.join(FOLDER, f"{SAFE_TITLE}.xlsx")

# ============================== CONFIG (optional) ==============================
# Manual overrides/additions on top of auto-detected trips (see module docstring).
# Reimbursed trips: (start_date, end_date, [extra merchant keywords seen during the trip])
TRIPS = [
    # (dt.date(2026, 6, 22), dt.date(2026, 6, 26), ["UBER", "AIRPORT"]),
]
TRIP_CITY_HINTS = []          # e.g. ["CHICAGO"] — matched anywhere, any date
REIMB_KEYWORDS = [
    # ("EMPLOYER NAME", "DIRECT-PAY"),   # all keywords in a tuple must match the same deposit
]

# Auto-detection tuning (usually fine to leave as-is):
TRIP_GAP_DAYS = 4          # away-from-home charges within this many days of each other join one trip
TRIP_MIN_TXNS = 2          # ...unless a cluster has fewer than this many charges...
TRIP_MIN_SPEND = 150.0     # ...and less than this much total spend, in which case it's ignored (too weak to call a trip)
TRIP_SMALL_CLUSTER_FLOOR = 50.0   # a 2+ charge cluster still needs at least this much spend
REIMB_WINDOW_DAYS = 45     # how many days after a trip ends to look for a matching reimbursement deposit
REIMB_MATCH_RANGE = (0.5, 1.5)    # candidate deposit must be within this range of the trip's cost
US_STATES = {"AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA","KS","KY","LA",
             "ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ","NM","NY","NC","ND","OH","OK",
             "OR","PA","RI","SC","SD","TN","TX","UT","VT","VA","WA","WV","WI","WY","DC"}
LOCATION_RE = re.compile(r"([A-Za-z][A-Za-z.'&\- ]{1,28}?)\s{2,}([A-Z]{2})\s*$")
# ================================================================================

PENDING_LABEL = "Pending / Uncategorized"
CAT_RENAME = {"Uncategorized": "Other", "Atm Fee": "Cash", "Food & Dining": "Restaurants"}
INCOME_CATS = ("Paycheck", "Interest Income", "State Tax", "Federal Tax", "Income", "Refund")
P2P_KEYWORDS = ["ZELLE", "VENMO", "CASH APP", "CASHAPP", "PAYPAL"]
HOUSING_HINT = re.compile(r"rent|mortgage", re.I)

CUR, CUR0, PCT = '"$"#,##0.00', '"$"#,##0', '0.0%'

# ------------------------- load files & remove repeated entries -------------------------
# Add as many CSVs as you have accounts — checking, savings, credit cards,
# brokerage, a second checking account, whatever. Each is auto-labeled from
# its filename: known keywords are matched first, otherwise the filename
# itself (cleaned up) becomes the account name. Name your files accordingly,
# e.g. "chase_credit.csv", "ally_savings.csv", "spouse_checking.csv".
ACCOUNT_HINTS = [("saving", "Savings"), ("checking", "Checking"), ("credit", "Credit Card"),
                 ("visa", "Credit Card"), ("mastercard", "Credit Card"), ("amex", "Credit Card"),
                 ("invest", "Investment"), ("brokerage", "Investment"), ("401k", "Retirement"),
                 ("ira", "Retirement")]

def merchant_key(desc):
    tokens = desc.split()
    while tokens and tokens[-1].isdigit():
        tokens.pop()
    return " ".join(tokens).strip().lower() or desc.lower()

def detect_account(path):
    base = os.path.splitext(os.path.basename(path))[0]
    low = base.lower()
    for hint, label in ACCOUNT_HINTS:
        if hint in low:
            return label
    cleaned = re.sub(r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\d*[-_]?\d*[-_]?\d*$", "", low, flags=re.I)
    cleaned = re.sub(r"[\d_\-]+$", "", cleaned)
    cleaned = re.sub(r"[_\-]+", " ", cleaned).strip()
    return cleaned.title() if cleaned else "Checking"

def load_file(path):
    acct = detect_account(path)
    out = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            if not r.get("Date"):
                continue
            out.append({"date": dt.date.fromisoformat(r["Date"]), "desc": r["Description"].strip(),
                        "orig": r["Original Description"].strip(), "cat": r["Category"].strip(),
                        "amt": float(r["Amount"]), "status": r["Status"].strip(), "acct": acct})
    return out

csv_files = sorted(glob.glob(os.path.join(FOLDER, "*.csv")))
if not csv_files:
    print("No CSV files found in this folder. Export your transactions here and run again.")
    ask("Press Enter to close...")
    sys.exit(1)
print("Reading:", ", ".join(os.path.basename(p) for p in csv_files))

per_key = defaultdict(dict)
for path in csv_files:
    for t in load_file(path):
        key = (t["acct"], t["date"], t["orig"], round(t["amt"], 2))
        per_key[key].setdefault(path, []).append(t)
rows = []
for key, byfile in per_key.items():
    best = max(byfile.values(), key=lambda rs: (sum(r["status"] == "Posted" for r in rs), len(rs)))
    rows.extend(best)
print(f"Transactions found (after removing repeats): {len(rows)}")

# ------------------------- auto-detect trips from location + spending pattern -------------------------
def extract_state(orig):
    m = LOCATION_RE.search(orig)
    if not m:
        return None
    st = m.group(2)
    return st if st in US_STATES else None

for t in rows:
    m = LOCATION_RE.search(t["orig"])
    t["city"] = m.group(1).strip().title() if m else None
    t["state"] = extract_state(t["orig"])
    t["trip"] = False
    t["reimb"] = False

state_counts = Counter(t["state"] for t in rows if t["state"])
home_state = state_counts.most_common(1)[0][0] if state_counts else None

# Online subscriptions (Spotify, Amazon, Whoop, Apple...) often bill from an
# out-of-state address every time — that's not travel. Two tells separate
# them from a real away-from-home purchase: (1) the description references
# the vendor's own web domain, since a physical merchant's location field
# never does, and (2) the charged amount barely varies between occurrences,
# since real travel spend (gas, food, hotels) does vary visit to visit.
DOMAIN_RE = re.compile(r"\.(COM|NET|ORG|CO|BIZ)\b", re.I)
away_amounts = defaultdict(list)
for t in rows:
    if t["state"] and t["state"] != home_state and t["amt"] < 0:
        away_amounts[merchant_key(t["desc"])].append(-t["amt"])

def is_subscription_like(desc, key):
    if DOMAIN_RE.search(desc):
        return True
    amts = away_amounts.get(key, [])
    if len(amts) < 3:
        return False
    mean = sum(amts) / len(amts)
    return mean > 0 and (max(amts) - min(amts)) <= max(2.0, 0.1 * mean)

# Rideshare/hotel/flight apps always bill from their own corporate address
# (e.g. Uber shows "uber.com...CA" no matter where the ride happened), which
# would otherwise get excluded by the domain check above. These categories
# are travel almost by definition, so they're always kept as real signal.
TRAVEL_CATEGORIES = {"Hotel", "Air Travel", "Rental Car & Taxi", "Travel"}

detected_trips = []   # for reporting: dicts with start/end/cost/locations/reimb_amt/reimb_date
if home_state:
    away = sorted([t for t in rows if t["state"] and t["state"] != home_state and t["amt"] < 0
                   and (t["cat"] in TRAVEL_CATEGORIES
                        or not is_subscription_like(t["orig"], merchant_key(t["desc"])))],
                  key=lambda t: t["date"])
    clusters, current = [], []
    for t in away:
        if current and (t["date"] - current[-1]["date"]).days > TRIP_GAP_DAYS:
            clusters.append(current); current = []
        current.append(t)
    if current:
        clusters.append(current)

    def cluster_cost(c): return sum(-x["amt"] for x in c)
    kept = [c for c in clusters
            if cluster_cost(c) >= TRIP_MIN_SPEND
            or (len(c) >= TRIP_MIN_TXNS and cluster_cost(c) >= TRIP_SMALL_CLUSTER_FLOOR)]

    income_candidates = [t for t in rows if t["amt"] > 0 and t["cat"] == "Income"]
    used_reimb_ids = set()
    for c in kept:
        for t in c:
            t["trip"] = True
        start, end = min(t["date"] for t in c), max(t["date"] for t in c)
        cost = cluster_cost(c)
        window_end = end + dt.timedelta(days=REIMB_WINDOW_DAYS)
        lo, hi = REIMB_MATCH_RANGE
        candidates = [t for t in income_candidates
                      if start <= t["date"] <= window_end and id(t) not in used_reimb_ids
                      and lo * cost <= t["amt"] <= hi * cost]
        reimb_amt, reimb_date = None, None
        if candidates:
            best = min(candidates, key=lambda t: abs(t["amt"] - cost))
            best["reimb"] = True
            used_reimb_ids.add(id(best))
            reimb_amt, reimb_date = best["amt"], best["date"]
        locs = sorted({t["state"] for t in c})
        detected_trips.append({"start": start, "end": end, "cost": cost, "count": len(c),
                                "locations": locs, "reimb_amt": reimb_amt, "reimb_date": reimb_date})

# ------------------------- classify -------------------------
for t in rows:
    up = t["orig"].upper()
    raw_cat = t["cat"]
    if raw_cat == "Category Pending":
        t["cat"] = PENDING_LABEL
    else:
        t["cat"] = CAT_RENAME.get(raw_cat, raw_cat)
    in_window = any(s <= t["date"] <= e and any(k in up for k in extras) for s, e, extras in TRIPS)
    if bool(TRIP_CITY_HINTS and any(h in up for h in TRIP_CITY_HINTS)) or in_window:
        t["trip"] = True
    if any(all(k in up for k in ks) for ks in REIMB_KEYWORDS):
        t["reimb"] = True
    is_p2p = any(k in up for k in P2P_KEYWORDS)
    t["internal"] = (raw_cat == "Transfer") and not is_p2p
    if is_p2p:
        t["cat"] = "Personal Payments"
        if t["amt"] > 0:
            raw_cat = "Income"
    t["income"] = (t["amt"] > 0 and raw_cat in INCOME_CATS
                   and not (t["reimb"] or t["internal"]))

# ------------------------- optional: AI categorization -------------------------
# Two ways to enable this - neither is required for the tool to work:
#   1. Ollama running locally (https://ollama.com) - free, no API key, no
#      internet needed once you've pulled a model. Preferred automatically
#      when detected, since it fits "share this with anyone" best.
#   2. An ANTHROPIC_API_KEY environment variable - uses the Claude API instead.
#      Costs a small amount per run and needs internet access.
# Neither set up -> this step is skipped entirely and nothing else changes.
AI_MODEL = "claude-opus-4-8"   # used only if Ollama isn't running; change to "claude-haiku-4-5" for cheaper/faster
OLLAMA_URL = "http://localhost:11434"
OLLAMA_MODEL = None            # None = auto-use whatever model you have installed; or set e.g. "llama3.2"

CATEGORIZE_PROMPT = ("For each merchant/description below, reply with a short, general spending "
                      "category (like Groceries, Restaurants, Gas, Shopping, Utilities). Reply with "
                      "ONLY a JSON object mapping each description to its category - no other text.\n\n")

def ollama_models():
    import urllib.request
    try:
        with urllib.request.urlopen(OLLAMA_URL + "/api/tags", timeout=1.5) as resp:
            data = json.loads(resp.read())
        return [m.get("name") for m in data.get("models", []) if m.get("name")]
    except Exception:
        return None

def ollama_ask(prompt, model):
    import urllib.request
    body = json.dumps({"model": model, "prompt": prompt, "stream": False, "format": "json"}).encode()
    req = urllib.request.Request(OLLAMA_URL + "/api/generate", data=body,
                                  headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=180) as resp:
        return json.loads(resp.read()).get("response", "")

def ai_categorize(rows):
    needs_help = sorted({t["desc"] for t in rows if t["cat"] in (PENDING_LABEL, "Other")})
    if not needs_help:
        return
    models = ollama_models()
    has_key = bool(os.environ.get("ANTHROPIC_API_KEY"))
    if not models and not has_key:
        return
    use_ollama = bool(models)
    who = f"a local Ollama model ({OLLAMA_MODEL or models[0]})" if use_ollama else "Claude (cloud API)"
    ans = ask(f"Found {len(needs_help)} uncategorized merchants. Use {who} to guess categories? [y/N] ").strip().lower()
    if ans != "y":
        return
    prompt = CATEGORIZE_PROMPT + "\n".join(needs_help)
    try:
        if use_ollama:
            text = ollama_ask(prompt, OLLAMA_MODEL or models[0])
        else:
            try:
                import anthropic
            except ImportError:
                import subprocess
                subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet", "anthropic"])
                import anthropic
            client = anthropic.Anthropic()
            response = client.messages.create(model=AI_MODEL, max_tokens=4096,
                                               messages=[{"role": "user", "content": prompt}])
            if response.stop_reason == "refusal":
                print("AI categorization was declined; continuing without it.")
                return
            text = next((b.text for b in response.content if b.type == "text"), "")
        mapping = json.loads(text)
        applied = 0
        for t in rows:
            if t["desc"] in mapping and t["cat"] in (PENDING_LABEL, "Other"):
                t["cat"] = mapping[t["desc"]]
                applied += 1
        print(f"AI categorized {applied} transactions across {len(mapping)} merchants.")
    except Exception as e:
        print(f"AI categorization failed ({e}); continuing without it.")

ai_categorize(rows)

START, END = min(t["date"] for t in rows), max(t["date"] for t in rows)
DAYS = (END - START).days + 1
MONTHS = max(DAYS / 30.437, 1 / 30.437)

trip_cost = -sum(t["amt"] for t in rows if t["trip"])
trip_reimb = sum(t["amt"] for t in rows if t["reimb"])
# Transfers between your own accounts: summing only the outgoing (negative)
# side across ALL accounts gives the total moved, without double-counting
# even if both the source and destination accounts were exported.
xfer_total = sum(-t["amt"] for t in rows if t["internal"] and t["amt"] < 0)
xfer_by_acct = defaultdict(float)
for t in rows:
    if t["internal"] and t["amt"] < 0:
        xfer_by_acct[t["acct"]] += -t["amt"]

def spend_rows(adjusted=True):
    return [t for t in rows
            if not (t["internal"] or t["income"] or t["reimb"])
            and not (adjusted and t["trip"])]

inc_total = sum(t["amt"] for t in rows if t["income"])
exp_adj = sum(-t["amt"] for t in spend_rows(True))
inc_m, exp_m = inc_total / MONTHS, exp_adj / MONTHS
net_m, xfer_m = inc_m - exp_m, xfer_total / MONTHS
accounts = sorted({t["acct"] for t in rows})

cat_tot = defaultdict(float)
for t in spend_rows(True):
    cat_tot[t["cat"]] += -t["amt"]
cat_sorted = sorted(cat_tot.items(), key=lambda kv: -kv[1])

def mkey(d): return d.strftime("%Y-%m")
months = sorted({mkey(t["date"]) for t in rows})
m_inc, m_exp_adj, m_xfer = defaultdict(float), defaultdict(float), defaultdict(float)
for t in rows:
    if t["income"]:
        m_inc[mkey(t["date"])] += t["amt"]
    if t["internal"] and t["amt"] < 0:
        m_xfer[mkey(t["date"])] += -t["amt"]
for t in spend_rows(True):
    m_exp_adj[mkey(t["date"])] += -t["amt"]

def wk(d): return d - dt.timedelta(days=d.weekday())
w_adj, w_raw, w_inc = defaultdict(float), defaultdict(float), defaultdict(float)
for t in spend_rows(True):
    w_adj[wk(t["date"])] += -t["amt"]
for t in spend_rows(False):
    w_raw[wk(t["date"])] += -t["amt"]
for t in rows:
    if t["income"]:
        w_inc[wk(t["date"])] += t["amt"]

# ---- generic recurring-merchant detection (no hardcoded vendor names) ----
rec_groups = defaultdict(list)
for t in spend_rows(True):
    if t["amt"] >= 0:
        continue
    rec_groups[merchant_key(t["desc"])].append(t)
recurring = [(k.title(), ts) for k, ts in rec_groups.items() if len(ts) >= 2]
recurring.sort(key=lambda kv: -sum(-t["amt"] for t in kv[1]))

lumpy = sorted([t for t in spend_rows(True) if -t["amt"] >= 100 and not HOUSING_HINT.search(t["cat"])],
               key=lambda x: x["amt"])
has_transfers = xfer_total > 0

# ============================ workbook (plain, simple styling) ============================
wb = Workbook()
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
HEADER_FILL = "D9D9D9"

def F(sz=10, b=False, i=False):
    return Font(name="Calibri", size=sz, bold=b, italic=i)

def fillc(h):
    return PatternFill("solid", fgColor=h)

def title(ws, text, sub, cols=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = F(16, True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    c2 = ws.cell(row=2, column=1, value=sub)
    c2.font = F(9, i=True)
    ws.row_dimensions[1].height = 22

def table(ws, r0, col0, headers, data, widths, fmts, total=None):
    for i, h in enumerate(headers):
        c = ws.cell(row=r0, column=col0 + i, value=h)
        c.font = F(10, True); c.fill = fillc(HEADER_FILL); c.border = BOX
        c.alignment = Alignment(horizontal="left", vertical="center")
    r = r0 + 1
    for row in data:
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=col0 + i - 1, value=v)
            c.font = F(10); c.border = BOX
            if i in fmts and isinstance(v, (int, float)):
                c.number_format = fmts[i]
        r += 1
    if total:
        for i, v in enumerate(total, 1):
            c = ws.cell(row=r, column=col0 + i - 1, value=v)
            c.font = F(10, True); c.border = BOX
            if i in fmts and isinstance(v, (int, float)):
                c.number_format = fmts[i]
        r += 1
    for i, wd in enumerate(widths):
        ws.column_dimensions[get_column_letter(col0 + i)].width = wd
    return r

# ---------- Dashboard ----------
ws = wb.active; ws.title = "Dashboard"
acct_note = f"{len(accounts)} accounts: {', '.join(accounts)}" if len(accounts) > 1 else f"{accounts[0]} account"
title(ws, budget_title, f"{START:%b %d, %Y} - {END:%b %d, %Y}  ({DAYS} days, ~{MONTHS:.1f} months)  |  {acct_note}", 10)

kpis = [
    ("Total income / mo", inc_m, "Paychecks, refunds, interest"),
    ("Total expenses / mo", exp_m, "Adjusted spend"),
    ("Surplus / mo", net_m, "Income minus expenses"),
]
if has_transfers:
    kpis.append(("Moved between accounts / mo", xfer_m, f"${xfer_total:,.0f} total across {len(accounts)} accounts"))
r = table(ws, 4, 1, ["METRIC", "AMOUNT", "NOTE"], [[l, v, s] for l, v, s in kpis], [30, 15, 48], {2: CUR0})

r += 1
ws.cell(row=r, column=1, value="Where the money goes").font = F(12, True)
r += 1
cat_r0 = r
cdata = [[c, v, v / exp_adj if exp_adj else 0] for c, v in cat_sorted]
last_cat_row = table(ws, cat_r0, 1, ["CATEGORY", "AMOUNT", "% OF SPEND"], cdata, [26, 14, 12],
                     {2: CUR0, 3: PCT}, total=["TOTAL", exp_adj, 1.0])

# Donut references the visible table directly (top slices) — no separate
# hidden staging area needed.
# Small categories get bucketed into "Other" for the chart only (the table
# above still lists every category) — too many slivers makes the pie
# unreadable. Staged in columns J/K, hidden, off to the side of the table.
topn = cat_sorted[:7]
other_amt = sum(v for _, v in cat_sorted[7:])
donut_rows = topn + ([("Other", other_amt)] if other_amt > 0 else [])
d0 = cat_r0
for i, (c, v) in enumerate(donut_rows):
    ws.cell(row=d0 + i, column=10, value=c)
    ws.cell(row=d0 + i, column=11, value=v)
d_last = d0 + len(donut_rows) - 1
donut = DoughnutChart()
donut.title = "Where the money goes"
donut.add_data(Reference(ws, min_col=11, min_row=d0, max_row=d_last), titles_from_data=False)
donut.set_categories(Reference(ws, min_col=10, min_row=d0, max_row=d_last))
donut.height, donut.width = 9, 11
donut.visible_cells_only = False
donut.series[0].dLbls = DataLabelList()
donut.series[0].dLbls.showSerName = False
donut.series[0].dLbls.showCatName = False
donut.series[0].dLbls.showVal = False
donut.series[0].dLbls.showLegendKey = False
donut.series[0].dLbls.showPercent = True
ws.add_chart(donut, f"E{cat_r0}")
for hide_col in ("J", "K"):
    ws.column_dimensions[hide_col].hidden = True

r2 = last_cat_row + 1
gap = net_m - xfer_m
if gap >= 0 or not has_transfers:
    banner = f"Income covers expenses with ${net_m:,.0f}/mo to spare."
else:
    banner = f"Bills covered with ${net_m:,.0f}/mo to spare, but your savings pace runs ${-gap:,.0f}/mo ahead of that surplus."
ws.cell(row=r2, column=1, value=banner).font = F(11, True)
r2 += 2

if detected_trips:
    ws.cell(row=r2, column=1, value=f"Auto-detected trips (home base: {home_state})").font = F(11, True); r2 += 1
    trip_data = []
    for d in detected_trips:
        span = f"{d['start']:%b %d} - {d['end']:%b %d}" if d['start'] != d['end'] else f"{d['start']:%b %d}"
        where = ", ".join(d["locations"])
        reimb_note = f"${d['reimb_amt']:,.0f} on {d['reimb_date']:%b %d}" if d["reimb_amt"] else "none found — counted as a real expense"
        trip_data.append([span, where, d["cost"], reimb_note])
    r2 = table(ws, r2, 1, ["DATES", "STATE(S)", "COST", "REIMBURSEMENT"], trip_data,
               [20, 14, 12, 34], {3: CUR}, total=["TOTAL", "", trip_cost, ""]) + 1
    net_oop = trip_cost - trip_reimb
    ws.cell(row=r2, column=1, value=f"Net out of pocket across all detected trips: ${net_oop:,.0f}").font = F(9, i=True)
    r2 += 2
elif trip_cost or trip_reimb:
    ws.cell(row=r2, column=1, value="Trip reconciliation").font = F(11, True); r2 += 1
    r2 = table(ws, r2, 1, ["ITEM", "AMOUNT"],
               [["Trip expenses fronted", -trip_cost], ["Reimbursement received", trip_reimb],
                ["Net out of pocket", trip_reimb - trip_cost]], [30, 14], {2: CUR}) + 1

ws.cell(row=r2, column=1, value="Notable one-time purchases (>= $100)").font = F(11, True); r2 += 1
table(ws, r2, 1, ["PURCHASE", "AMOUNT"],
      [[f"{t['date']:%b %d} - {t['desc']} ({t['cat']})", -t["amt"]] for t in lumpy][:14],
      [50, 14], {2: CUR})

# ---------- Monthly ----------
ws = wb.create_sheet("Monthly")
title(ws, "Monthly Cash Flow", "Adjusted = detected/configured trip removed. Transfers between your own accounts shown separately.", 6)
data = [[dt.date.fromisoformat(k + "-01").strftime("%b %Y"), m_inc[k], m_exp_adj[k],
         m_inc[k] - m_exp_adj[k], m_xfer[k]] for k in months]
r = table(ws, 4, 1, ["MONTH", "INCOME", "EXPENSES", "NET", "TRANSFERRED OUT"], data,
          [12, 14, 14, 13, 16], {2: CUR0, 3: CUR0, 4: CUR0, 5: CUR0},
          total=["TOTAL", inc_total, exp_adj, inc_total - exp_adj, xfer_total])
n = len(months)
ch = BarChart(); ch.type = "col"; ch.title = "Income vs expenses by month"
ch.add_data(Reference(ws, min_col=2, min_row=4, max_row=4 + n), titles_from_data=True)
ch.add_data(Reference(ws, min_col=3, min_row=4, max_row=4 + n), titles_from_data=True)
ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4 + n))
ch.gapWidth = 60; ch.y_axis.numFmt = CUR0
ch.height, ch.width = 9.5, 21
ws.add_chart(ch, f"A{r + 2}")
ch2 = LineChart(); ch2.title = "Net cash flow by month"
ch2.add_data(Reference(ws, min_col=4, min_row=4, max_row=4 + n), titles_from_data=True)
ch2.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4 + n))
ch2.y_axis.numFmt = CUR0
ch2.height, ch2.width = 9.5, 21
ws.add_chart(ch2, f"A{r + 22}")

# ---------- Categories ----------
ws = wb.create_sheet("Categories")
title(ws, "Where the money goes", "Adjusted spend by category, full period.", 6)
table(ws, 4, 1, ["CATEGORY", "TOTAL", "PER MONTH", "% OF SPEND"],
      [[c, v, v / MONTHS, v / exp_adj if exp_adj else 0] for c, v in cat_sorted], [26, 14, 13, 12],
      {2: CUR, 3: CUR, 4: PCT}, total=["TOTAL", exp_adj, exp_m, 1.0])
# Chart shows the top 12 categories + "Other" (bucketed, staged in hidden
# columns J/K) - the full list is already in the table above; cramming all
# ~25+ categories into one chart is what caused the label overlap.
bar_top = cat_sorted[:12]
bar_other = sum(v for _, v in cat_sorted[12:])
bar_rows = bar_top + ([("Other", bar_other)] if bar_other > 0 else [])
b0 = 5
for i, (c, v) in enumerate(reversed(bar_rows)):
    ws.cell(row=b0 + i, column=10, value=c)
    ws.cell(row=b0 + i, column=11, value=v)
b_last = b0 + len(bar_rows) - 1
ch = BarChart(); ch.type = "bar"; ch.title = "Spend by category (top 12 + Other)"
ch.add_data(Reference(ws, min_col=11, min_row=b0, max_row=b_last), titles_from_data=False)
ch.set_categories(Reference(ws, min_col=10, min_row=b0, max_row=b_last))
ch.legend = None; ch.x_axis.numFmt = CUR0
ch.visible_cells_only = False
ch.height, ch.width = 13.5, 20
ws.add_chart(ch, "G4")
for cn in ("J", "K"):
    ws.column_dimensions[cn].hidden = True

# ---------- Trends ----------
ws = wb.create_sheet("Trends")
title(ws, "Spending trends", "Weekly spend (Mon-start weeks) and cumulative surplus.", 6)
allw = sorted(set(w_raw) | set(w_adj) | set(w_inc))
data, cum = [], 0.0
for wkk in allw:
    cum += w_inc.get(wkk, 0) - w_adj.get(wkk, 0)
    data.append([wkk.strftime("%b %d"), w_raw.get(wkk, 0.0), w_adj.get(wkk, 0.0), cum])
table(ws, 4, 1, ["WEEK OF", "SPEND (RAW)", "SPEND (ADJ)", "CUMULATIVE SURPLUS"], data,
      [11, 13, 13, 18], {2: CUR0, 3: CUR0, 4: CUR0})
nw = len(data)
# tickLblSkip thins out the week labels along the x-axis - with 20+ weeks,
# showing every single one crowds and overlaps the text.
label_skip = max(1, nw // 12)
ch = LineChart(); ch.title = "Weekly spending: raw vs adjusted"
ch.add_data(Reference(ws, min_col=2, min_row=4, max_col=3, max_row=4 + nw), titles_from_data=True)
ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4 + nw))
ch.y_axis.numFmt = CUR0
ch.x_axis.tickLblSkip = label_skip
ch.height, ch.width = 9.5, 22
ws.add_chart(ch, "F4")
ch2 = LineChart(); ch2.title = "Cumulative surplus"
ch2.add_data(Reference(ws, min_col=4, min_row=4, max_row=4 + nw), titles_from_data=True)
ch2.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4 + nw))
ch2.y_axis.numFmt = CUR0
ch2.x_axis.tickLblSkip = label_skip
ch2.height, ch2.width = 9.5, 22
ws.add_chart(ch2, "F24")

# ---------- Transfers (only if present) ----------
if has_transfers:
    ws = wb.create_sheet("Transfers")
    title(ws, "Transfers between your own accounts", f"Across {len(accounts)} detected accounts: {', '.join(accounts)}. Not counted as spending.", 6)
    mdata = [[dt.date.fromisoformat(k + "-01").strftime("%b %Y"), m_xfer[k]] for k in months]
    r = table(ws, 4, 1, ["MONTH", "TRANSFERRED OUT (all accounts)"], mdata, [16, 24], {2: CUR0},
              total=["TOTAL", xfer_total])
    ch = BarChart(); ch.type = "col"; ch.title = "Transfers by month"
    ch.add_data(Reference(ws, min_col=2, min_row=4, max_row=4 + len(mdata)), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4 + len(mdata)))
    ch.gapWidth = 60; ch.y_axis.numFmt = CUR0; ch.legend = None
    ch.height, ch.width = 9.5, 18
    ws.add_chart(ch, "D4")

    acct_data = sorted(xfer_by_acct.items(), key=lambda kv: -kv[1])
    table(ws, r + 2, 1, ["ACCOUNT (money left FROM here)", "TOTAL SENT OUT"],
          [[a, v] for a, v in acct_data], [30, 16], {2: CUR0})

# ---------- Recurring ----------
ws = wb.create_sheet("Recurring")
title(ws, "Recurring charges (auto-detected)", "Merchants seen 2+ times. 'Est. monthly' = total / months elapsed. Review before trusting fully.", 6)
data, sub_total = [], 0.0
for label, ts in recurring[:25]:
    tot = sum(-t["amt"] for t in ts); sub_total += tot
    data.append([label, len(ts), tot, tot / MONTHS])
table(ws, 4, 1, ["ITEM", "HITS", "TOTAL", "EST. MONTHLY"], data, [30, 7, 14, 14],
      {3: CUR, 4: CUR}, total=["TOTAL (shown)", "", sub_total, sub_total / MONTHS])

# ---------- Transactions ----------
ws = wb.create_sheet("Transactions")
title(ws, "All transactions (repeats removed)", "Flags: Trip / Reimb apply automatically, or from the optional CONFIG block. Internal = between your own accounts.", 7)
hdr = ["DATE", "ACCOUNT", "DESCRIPTION", "CATEGORY", "AMOUNT", "STATUS", "FLAGS"]
for i, h in enumerate(hdr, 1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = F(10, True); c.fill = fillc(HEADER_FILL); c.border = BOX
r = 5
for t in sorted(rows, key=lambda x: (x["date"], x["acct"]), reverse=True):
    flags = " ".join(f for f, b in [("Trip", t["trip"]), ("Reimb", t["reimb"]), ("Internal", t["internal"])] if b)
    for i, v in enumerate([t["date"], t["acct"], t["desc"], t["cat"], t["amt"], t["status"], flags], 1):
        c = ws.cell(row=r, column=i, value=v); c.font = F(10); c.border = BOX
    ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=5).number_format = CUR
    r += 1
ws.auto_filter.ref = f"A4:G{r - 1}"
ws.freeze_panes = "A5"
for i, wd in enumerate([11, 10, 36, 24, 12, 9, 16], 1):
    ws.column_dimensions[get_column_letter(i)].width = wd

# ------------------------- save (retry if locked) -------------------------
for attempt in range(1, 6):
    try:
        wb.save(OUT)
        break
    except PermissionError:
        print(f"'{os.path.basename(OUT)}' seems to be open in Excel. Close it and I'll retry in 5s... ({attempt}/5)")
        time.sleep(5)
else:
    print(f"Could not save — please close '{os.path.basename(OUT)}' and run this again.")
    ask("Press Enter to close...")
    sys.exit(1)

print(f"\nWindow: {START} - {END} (~{MONTHS:.1f} months)")
print(f"Income ${inc_m:,.0f}/mo | Expenses ${exp_m:,.0f}/mo | Surplus ${net_m:,.0f}/mo")
print(f"Saved: {OUT}")
try:
    os.startfile(OUT)
except Exception:
    pass
ask("Done! Press Enter to close this window...")

